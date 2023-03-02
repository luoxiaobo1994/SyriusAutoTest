import datetime
import os
import sys

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from com.syrius.util.commonUtil import util
from com.syrius.util.sshUtil import *

logger.remove()  # 删去import logger之后自动产生的handler，不删除的话会出现重复输出的现象
logger.add(sys.stderr, level="DEBUG")  # level=TRACE,DEBUG,INFO,SUCCESS,WARNING,ERROR,CRITICAL
# 使用脚本测试时，建议使用ERROR，调试脚本时使用DEBUG

"""
参考文档：https://www.jb51.net/article/211271.htm
"""


def save_excel_file(wb, robot_name):
    """
    保存文件
    :param wb:
    :return:
    """
    file_name = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S.%f")[:-3]
    fileDir = file_name[:file_name.find("_")]
    file_path = f"{util().get_project_path()}/file/userRight/{fileDir}"
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    wb.save(f"{file_path}/userRight_{robot_name}_{file_name}.xlsx")


def check_app_process(work_book, sheet_name):
    """
    应用程序进程权限检查
    :param work_book:
    :return:
    """
    sheet = work_book[sheet_name]
    row1 = 1
    for one_row in sheet.rows:
        # if str(one_row[1].value).isascii():  # 去除首行汉字
        if not row1 == 1:
            # print(one_row[1].value)
            result = ssh_app_process(one_row[1].value)
            one_row[4].value = result[0]  # 实际用户赋值，从0计数，第4列
            one_row[5].value = result[1]  # 实际用户组赋值
            user_flag = one_row[2].value.strip() == result[0].strip()
            group_flag = one_row[3].value.strip() == result[0].strip()
            if user_flag and group_flag:

                one_row[6].value = 'pass'  # 最终结果赋值
                logger.info(f"本次判定的结果为：【{one_row[6].value}】")
            else:
                one_row[6].value = 'false'
                one_row[6].fill = PatternFill(fill_type="solid", fgColor="00FF0000")
                logger.error(f"本次判定的结果为：【{one_row[6].value}】，服务为：【{one_row[1].value}】")
        row1 += 1  # 每次循环之后开始迭代+1行进行计数


def ssh_app_process(param):
    """
    应用程序进程的查询结果
    :param command:获取版本号的指令
    :return:返回版本信息
    """
    # 比较全的指令参数为：
    # ps axo user:20,pid,pcpu,pmem,vsz,rss,tty,stat,start,time,cmd 2>1 |grep scan
    command = f"ps -eo user:20,group:20,cmd | grep -v color=auto |grep -v grep | grep -v run.sh | grep {param}"
    robotSystemInfo = ssh_exe_cmd(command)
    result = ['', '', '']
    if len(robotSystemInfo.strip()) != 0:
        result = re.split('[ ]{1,}', robotSystemInfo)  # 1个以上的空格进行分割
    logger.info(f"返回的数组为：{result}")
    return result


def ssh_app_file(command=None):
    """
    应用程序文件权限检查
    :param command:获取版本号的指令
    :return:返回版本信息
    """
    if command is None:
        command = f"pwd"
    robotSystemInfo = ssh_exe_cmd(command)
    result = ['', '', '', '']
    if len(robotSystemInfo.strip()) != 0:
        result = re.split('[ ]{1,}', robotSystemInfo)  # 1个以上的空格进行分割
    logger.info(f"返回的数组为：{result}")
    return result


def ssh_app_dir(command=None):
    """
    应用程序文件权限检查
    :param command:获取版本号的指令
    :return:返回版本信息
    """
    if command is None:
        command = f"pwd"
    robotSystemInfo = ssh_exe_cmd(command)
    result = ['', '', '', '']
    if len(robotSystemInfo.strip()) != 0:
        first_robotSystemInfo = re.split("\\n", robotSystemInfo)[1]
        result = re.split('[ ]{1,}', first_robotSystemInfo)  # 1个以上的空格进行分割
    logger.info(f"返回的数组为：{result}")
    return result


def change_num(param):
    """
    用户权限字符串转化为数字
    :param param:
    :return:
    """
    param = param[1:]
    paramArray = re.findall(r'.{3}', param)
    param_value = ''
    for one_param in paramArray:
        if one_param == '--x':
            param_value += '1'
        elif one_param == '-w-':
            param_value += '2'
        elif one_param == 'r--':
            param_value += '4'
        elif one_param == '-wx':
            param_value += '3'
        elif one_param == 'r-x':
            param_value += '5'
        elif one_param == 'rw-':
            param_value += '6'
        elif one_param == 'rwx':
            param_value += '7'
        elif one_param == '---':
            param_value += '0'
    logger.debug(f"将字符串【{param}】转化为【{param_value}】")
    return param_value


def check_app_file(work_book, sheet_name):
    """
    文件权限检查
    :param work_book:
    :return:
    """
    sheet = work_book[sheet_name]
    row1 = 1
    for one_row in sheet.rows:
        if not row1 == 1:
            result = ssh_app_file(one_row[9].value)  # 提供查询指令
            one_row[5].value = result[2]  # 实际用户赋值，从0计数，第4列
            one_row[6].value = result[3]  # 实际用户组赋值
            num = change_num(result[0].strip())
            one_row[7].value = num  # 实际权限赋值
            user_flag = one_row[2].value.strip() == result[2].strip()
            group_flag = one_row[3].value.strip() == result[3].strip()
            permisssion_flag = str(one_row[4].value).strip() == num
            if user_flag and group_flag and permisssion_flag:
                one_row[8].value = 'pass'  # 最终结果赋值
                logger.info(f"本次判定的结果为：【{one_row[8].value}】")
            else:
                one_row[8].value = 'false'
                one_row[8].fill = PatternFill(fill_type="solid", fgColor="00FF0000")
                logger.error(f"本次判定的结果为：【{one_row[8].value}】，查询指令为：【{one_row[9].value}】")
        row1 += 1  # 每次循环之后开始迭代+1行进行计数


def check_app_dir(work_book, sheet_name):
    """
    文件夹权限检查
    :param wb:
    :param param:
    :return:
    """
    sheet = work_book[sheet_name]
    row1 = 1
    for one_row in sheet.rows:
        if not row1 == 1:
            result = ssh_app_dir(f"ls -al {one_row[1].value}")  # 提供查询指令
            one_row[5].value = result[2]  # 实际用户赋值，从0计数，第4列
            one_row[6].value = result[3]  # 实际用户组赋值
            num = change_num(result[0].strip())
            one_row[7].value = num  # 实际权限赋值
            user_flag = one_row[2].value.strip() == result[2].strip()
            group_flag = one_row[3].value.strip() == result[3].strip()
            permisssion_flag = str(one_row[4].value).strip() == num
            if user_flag and group_flag and permisssion_flag:
                one_row[8].value = 'pass'  # 最终结果赋值
                logger.info(f"本次判定的结果为：【{one_row[8].value}】")
            else:
                one_row[8].value = 'false'
                one_row[8].fill = PatternFill(fill_type="solid", fgColor="00FF0000")
                logger.error(f"本次判定的结果为：【{one_row[8].value}】，查询指令为：【{one_row[9].value}】")
        row1 += 1  # 每次循环之后开始迭代+1行进行计数


def handle_excel(file_path, robot_name):
    """
    处理excel中的数据
    :param file_path:
    :return:
    """
    wb = load_workbook(file_path)
    check_app_process(wb, "app_process")  # 应用进程权限检查
    check_app_file(wb, "app_file")
    check_app_dir(wb, "app_dir")
    check_app_dir(wb, "data_dir")
    check_app_dir(wb, "log_dir")

    save_excel_file(wb, robot_name)


def main(addr, robot_name):
    '''  登录信息拼接，开启链接 '''
    address = addr.split(":")
    ssh_login(address[0], address[1], "developer", "developer")

    ''' 获取对应的信息 '''
    file_path = f"{util().get_project_path()}file\\userRight\\userRight_source.xlsx"

    handle_excel(file_path, robot_name)

    ''' 关闭链接 '''
    closeSSH()


if __name__ == '__main__':
    robotMap = {
        '雷龙-齐达内': '10.2.8.65:22',
        '雷龙-内马尔': '10.2.8.57:22',
        '雷龙-苏亚雷斯': '10.2.9.181:22',
        '雷龙-C罗': '10.2.8.118:22',
        '梁龙-佐助': '10.2.8.103:22',
        '梁龙-鸣人': '10.2.8.103:22',
        '梁龙2-罗宾': '10.2.8.77:22',
        '梁龙2-罗宾2': '10.2.9.112:22',
        '梁龙2-索隆': '10.2.9.82:22',
        '电子平台1-无外网': '10.2.8.193:22',
        '电子平台1-有外网': '10.2.16.200:9542'
    }
    robot_name = "梁龙2-罗宾2"
    main(robotMap[robot_name], robot_name)
